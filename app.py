from flask import Flask, render_template, request, send_file, url_for
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from email_service import email_service
from digil_test_service import digil_test_service
from flask_socketio import SocketIO, emit
from monitoring_service import AlarmMonitor
import os
import tempfile
import re
import shutil
import webbrowser
import threading
import time
import sys
import requests
import json
import urllib.parse
import zipfile
from pathlib import Path

# Gestione percorsi per PyInstaller
if getattr(sys, 'frozen', False):
    # Se è un exe compilato
    application_path = sys._MEIPASS
    running_as_exe = True
else:
    # Se è uno script Python normale
    application_path = os.path.dirname(os.path.abspath(__file__))
    running_as_exe = False

# Cambia directory di lavoro
os.chdir(os.path.dirname(sys.executable) if running_as_exe else application_path)

# Configurazione Flask
app = Flask(__name__, 
            template_folder=os.path.join(application_path, 'templates'),
            static_folder=os.path.join(application_path, 'static'))

socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
alarm_monitor = AlarmMonitor(socketio)

# Variabili globali per gestire il token
current_token = None
token_expires_at = None

def get_auth_token():
    """Ottiene un nuovo token di autenticazione"""
    global current_token, token_expires_at
    
    url = "https://rh-sso.apps.clusterzac.opencs.servizi.prv/auth/realms/DigilV2/protocol/openid-connect/token"
    
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Cookie': '86ff86b2f35123846fde75fccaac9811=f7ac6d867938d069d19133ee2b4f177b'
    }
    
    data = {
        'grant_type': 'client_credentials',
        'client_id': 'application', 
        'client_secret': 'q3pH03oAvt9io1K1rJ9GHVVRcmAEf55x'
    }
    
    try:
        response = requests.post(url, headers=headers, data=data, verify=False)
        response.raise_for_status()
        
        token_data = response.json()
        current_token = token_data['access_token']
        
        # Il token scade dopo 300 secondi, impostiamo scadenza a 290s per sicurezza
        token_expires_at = datetime.now() + timedelta(seconds=290)
        
        print(f"✅ Token ottenuto, scade alle: {token_expires_at.strftime('%H:%M:%S')}")
        return current_token
        
    except Exception as e:
        print(f"❌ Errore ottenimento token: {e}")
        raise Exception(f"Impossibile ottenere il token di autenticazione: {e}")

def is_token_valid():
    """Verifica se il token corrente è ancora valido"""
    global current_token, token_expires_at
    
    if not current_token or not token_expires_at:
        return False
    
    # Controlla se il token è scaduto (con margine di 10 secondi)
    return datetime.now() < (token_expires_at - timedelta(seconds=10))

def get_valid_token():
    """Ottiene un token valido, rinnovandolo se necessario"""
    if not is_token_valid():
        print("🔄 Token scaduto o mancante, ottengo nuovo token...")
        return get_auth_token()
    else:
        print("✅ Token ancora valido")
        return current_token

def get_device_id_from_api(device_name):
    """Ottiene il vero Device ID tramite API"""
    try:
        # Ottieni un token valido
        token = get_valid_token()
        
        # URL encode del nome dispositivo
        encoded_name = urllib.parse.quote(device_name)
        url = f"https://digil-back-end-onesait.servizi.prv/api/v1/digils?name={encoded_name}"
        
        headers = {
            'Accept': 'application/json',
            'Authorization': f'Bearer {token}',
            'Cookie': '7d1097d461d8e1d826eafd90fa29677c=6d499b21fdd884f7ee7979268d91e421'
        }
        
        print(f"🔍 Ricerca device: {device_name}")
        response = requests.get(url, headers=headers, verify=False)
        response.raise_for_status()
        
        data = response.json()
        
        # Estrae l'ID dal primo elemento del content
        if data.get('content') and len(data['content']) > 0:
            device_id = data['content'][0]['id']
            print(f"✅ Device ID trovato: {device_id}")
            return device_id
        else:
            raise Exception(f"Dispositivo non trovato: {device_name}")
            
    except Exception as e:
        print(f"❌ Errore ricerca device: {e}")
        raise Exception(f"Impossibile ottenere Device ID per {device_name}: {e}")

def transform_device_id(device_id):
    """Trasforma 1:1:2:16:22:DIGIL_MRN_0299 in 1121622_0299 (FALLBACK)"""
    # Rimuovi i ":"
    no_colons = device_id.replace(":", "")
    # Estrai solo i numeri
    numbers = ''.join(filter(str.isdigit, no_colons))
    # Prendi i primi numeri e gli ultimi 4 con "_"
    if len(numbers) >= 4:
        main_part = numbers[:-4]
        last_four = numbers[-4:]
        return f"{main_part}_{last_four}"
    return numbers

def transform_device_id_new(device_name):
    """Nuova funzione che ottiene il Device ID dalle API invece di trasformarlo"""
    try:
        # Ottiene il vero Device ID dalle API
        api_device_id = get_device_id_from_api(device_name)
        return api_device_id
    except Exception as e:
        print(f"⚠️ Errore API, uso trasformazione locale: {e}")
        # Fallback alla vecchia trasformazione se le API non funzionano
        return transform_device_id(device_name)

def format_date_for_sheet(date_str, time_str):
    """Formatta data per i sheet"""
    # Converte da YYYY-MM-DD a DD/MM/YYYY
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_date = date_obj.strftime("%d/%m/%Y")
    return f"{formatted_date} - {time_str}"

def format_date_for_api(start_date, start_time, end_date, end_time):
    """Formatta date per le API nel formato ISO"""
    # Aggiungi i secondi se mancano
    if len(start_time) == 5:  # HH:MM
        start_time += ":00"   # HH:MM:SS
    if len(end_time) == 5:    # HH:MM  
        end_time += ":00"     # HH:MM:SS
    
    start_dt = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M:%S")
    
    start_iso = start_dt.strftime("%Y-%m-%dT%H:%M:%S.000000000Z")
    end_iso = end_dt.strftime("%Y-%m-%dT%H:%M:%S.999999999Z")
    
    return start_iso, end_iso

def get_sensor_metrics(num_sensors):
    """Restituisce le metriche per il numero di sensori"""
    metrics = {
        3: {"range_metriche": 16, "allarme_metriche": 10, "allarmi": 4},
        6: {"range_metriche": 25, "allarme_metriche": 13, "allarmi": 7},
        12: {"range_metriche": 43, "allarme_metriche": 19, "allarmi": 13}
    }
    return metrics.get(num_sensors, metrics[3])

def update_common_parameters(ws, data):
    """Aggiorna i parametri comuni nel worksheet - RICERCA GENERICA"""
    print(f"=== DEBUG SHEET: {ws.title} ===")
    
    # Prima aggiorna il titolo
    try:
        if ws['A1'].value and "Test Report -" in str(ws['A1'].value):
            date_obj = datetime.strptime(data['start_date'], "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d/%m/%Y")
            old_value = ws['A1'].value
            ws['A1'].value = f"Test Report - {formatted_date}"
            print(f"A1 - Da: '{old_value}' A: '{ws['A1'].value}'")
    except:
        pass
    
    # Scansione completa con pattern più generici
    for row_num in range(1, 25):  # Aumentiamo il range
        for col_num in range(1, 10):
            try:
                cell = ws.cell(row=row_num, column=col_num)
                
                if cell.value and isinstance(cell.value, str):
                    
                    # Cerca date in formato DD/MM/YYYY - HH:MM:SS (per Inizio/Termine Test)
                    if "05/08/2025 -" in cell.value or "05/08/2025-" in cell.value:
                        if "13:28" in cell.value or "09:00" in cell.value:  # Inizio Test
                            print(f"TROVATO Inizio Test in {cell.coordinate}: '{cell.value}'")
                            old_value = cell.value
                            cell.value = data['start_test']
                            print(f"AGGIORNATO {cell.coordinate}: '{old_value}' → '{cell.value}'")
                        elif "14:02" in cell.value or "18:00" in cell.value:  # Termine Test  
                            print(f"TROVATO Termine Test in {cell.coordinate}: '{cell.value}'")
                            old_value = cell.value
                            cell.value = data['end_test']
                            print(f"AGGIORNATO {cell.coordinate}: '{old_value}' → '{cell.value}'")
                    
                    # Cerca specificamente "MII" come vendor (non toccare Topics Involved)
                    elif cell.value.strip() == "MII" and row_num > 5:  # Evita headers
                        print(f"TROVATO Vendor MII in {cell.coordinate}: '{cell.value}'")
                        old_value = cell.value
                        cell.value = data['vendor']
                        print(f"AGGIORNATO {cell.coordinate}: '{old_value}' → '{cell.value}'")
                    
                    # Cerca Device ID con pattern specifico
                    elif cell.value.startswith("1:1:2:16:22:DIGIL_MRN_") or cell.value.startswith("1:1:2:15:21:DIGIL_IND_"):
                        print(f"TROVATO Device ID in {cell.coordinate}: '{cell.value}'")
                        old_value = cell.value
                        cell.value = data['device_id']
                        print(f"AGGIORNATO {cell.coordinate}: '{old_value}' → '{cell.value}'")
                        
            except Exception as e:
                continue
    
    print(f"=== FINE DEBUG SHEET: {ws.title} ===\n")

def update_downlink_parameters(ws, transformed_device_id, data):
    """Aggiorna i parametri dello sheet Downlink - VERSIONE CORRETTA"""
    start_iso, end_iso = format_date_for_api(data['start_date'], data['start_time'], 
                                            data['end_date'], data['end_time'])
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Sostituisci SOLO gli ID dopo {deviceId}= mantenendo il formato
                if "{deviceId}=" in cell.value:
                    # Pattern: {deviceId}=1121622_0299 -> {deviceId}=API_DEVICE_ID
                    cell.value = re.sub(
                        r'\{deviceId\}=[^,\s}]+', 
                        f'{{deviceId}}={transformed_device_id}', 
                        cell.value
                    )
                
                # Aggiorna le date ISO
                if "startDate=" in cell.value:
                    cell.value = re.sub(
                        r'startDate=\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{9}Z',
                        f'startDate={start_iso}',
                        cell.value
                    )
                if "endDate=" in cell.value:
                    cell.value = re.sub(
                        r'endDate=\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{9}Z',
                        f'endDate={end_iso}',
                        cell.value
                    )

def create_excel_report(data):
    """Crea il report Excel compilato basandosi sui template reali"""
    
    # Gestisce percorsi sia per sviluppo che per exe
    if getattr(sys, 'frozen', False):
        # Se è un exe, usa il percorso temporaneo di PyInstaller
        base_path = sys._MEIPASS
    else:
        # Se è in sviluppo, usa il percorso normale
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    # Carica il template appropriato
    template_path = os.path.join(base_path, "templates_excel", f"esempio_{data['num_sensors']:02d}.xlsx")
    
    print(f"Cercando template in: {template_path}")  # Debug
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template non trovato: {template_path}")
    
    wb = load_workbook(template_path)
    
    # NOVITÀ: Ottiene il Device ID dalle API invece di trasformarlo
    print("🔄 Ottenimento Device ID dalle API...")
    transformed_device_id = transform_device_id_new(data['device_id'])
    
    start_formatted = format_date_for_sheet(data['start_date'], data['start_time'])
    end_formatted = format_date_for_sheet(data['end_date'], data['end_time'])
    
    common_data = {
        'start_test': start_formatted,
        'end_test': end_formatted,
        'vendor': data['vendor'],
        'device_id': data['device_id'],
        'start_date': data['start_date'],
        'start_time': data['start_time'],
        'end_date': data['end_date'],
        'end_time': data['end_time']
    }
    
    # Aggiorna ogni sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Aggiorna parametri comuni
        update_common_parameters(ws, common_data)
        
        # Aggiornamenti specifici per Downlink
        if "Downlink" in sheet_name:
            update_downlink_parameters(ws, transformed_device_id, common_data)
    
    return wb

@app.route('/')
def index():
    return render_template('index.html')

# Aggiungi questi import all'inizio del file app.py (dopo gli altri import)
import zipfile
from pathlib import Path

@app.route('/generate', methods=['POST'])
def generate_report():
    try:
        # Raccolta dati dal form
        data = {
            'num_sensors': int(request.form['num_sensors']),
            'start_date': request.form['start_date'],
            'start_time': request.form['start_time'],
            'end_date': request.form['end_date'],
            'end_time': request.form['end_time'],
            'vendor': request.form['vendor'],
            'device_id': request.form['device_id']
        }
        
        # Controlla se inviare email
        send_email = request.form.get('send_email') == 'on'
        custom_email = request.form.get('custom_email', '').strip()
        
        # Genera Excel
        wb = create_excel_report(data)
        
        # Estrai solo le ultime 4 cifre del Device ID per il nome
        # es. da "1:1:2:16:22:DIGIL_MRN_0299" prendi "0299"
        device_digits = ''.join(filter(str.isdigit, data['device_id']))
        device_short = device_digits[-4:] if len(device_digits) >= 4 else device_digits
        
        # Formatta la data per il nome file (DD-MM-YYYY)
        date_obj = datetime.strptime(data['start_date'], "%Y-%m-%d")
        date_formatted = date_obj.strftime("%d-%m-%Y")
        
        # Pulisci il vendor name per il filename (rimuovi caratteri speciali)
        vendor_clean = data['vendor'].replace('/', '-').replace('\\', '-')
        
        # Nuovo naming convention per il ZIP
        # "Report MII - 02-09-2025 - 464.zip"
        zip_filename = f"Report {vendor_clean} - {date_formatted} - {device_short}.zip"
        
        # Nome del file Excel dentro il ZIP (manteniamo il formato originale)
        datetime_start = datetime.strptime(f"{data['start_date']} {data['start_time']}", "%Y-%m-%d %H:%M")
        excel_filename = f"Report_Device_Fabbrica_generale_{datetime_start.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        
        # Crea cartella output se non esiste
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)
        
        # Crea sottocartella per vendor se non esiste
        vendor_dir = output_dir / vendor_clean
        vendor_dir.mkdir(exist_ok=True)
        
        # Percorso completo del file ZIP
        zip_path = vendor_dir / zip_filename
        
        # Salva prima il file Excel in temp
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_excel.name)
        temp_excel.close()
        
        # Crea il file ZIP e aggiungi l'Excel
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(temp_excel.name, excel_filename)
        
        # Rimuovi il file temporaneo
        os.unlink(temp_excel.name)
        
        print(f"✅ Report salvato in: {zip_path}")
        
        # Invia email se richiesto
        email_result = None
        if send_email:
            print("📧 Tentativo invio email...")
            
            try:
                # Importa il servizio email
                from email_service import email_service
                
                # Se c'è un'email custom, usa quella
                if custom_email:
                    custom_recipients = {'to': [custom_email], 'cc': []}
                    success, message = email_service.send_report_email(
                        zip_path, data['vendor'], data['device_id'], 
                        date_formatted, custom_recipients
                    )
                else:
                    success, message = email_service.send_report_email(
                        zip_path, data['vendor'], data['device_id'], date_formatted
                    )
                
                email_result = {'success': success, 'message': message}
                print(f"📧 {message}")
                
            except Exception as email_error:
                email_result = {
                    'success': False, 
                    'message': f"Errore invio email: {str(email_error)}"
                }
                print(f"❌ Errore invio email: {str(email_error)}")
        
        # Ritorna messaggio di successo con info email
        return render_template('success.html', 
                             zip_filename=zip_filename,
                             zip_path=str(zip_path),
                             vendor=data['vendor'],
                             device_id=data['device_id'],
                             date_formatted=date_formatted,
                             email_result=email_result)
    
    except Exception as e:
        return render_template('error.html', error_message=str(e))

@app.route('/preview', methods=['POST'])
def preview_report():
    try:
        # Validazione campi obbligatori
        required_fields = {
            'num_sensors': 'Numero Sensori',
            'device_id': 'Device ID', 
            'vendor': 'Vendor',
            'start_date': 'Data Inizio',
            'start_time': 'Ora Inizio'
        }
        
        missing_fields = []
        for field, label in required_fields.items():
            if not request.form.get(field):
                missing_fields.append(label)
        
        if missing_fields:
            fields_str = ", ".join(missing_fields)
            return f"""
            <div class="alert alert-warning">
                <h5>⚠️ Campi Mancanti</h5>
                <p>Per favore compila tutti i campi obbligatori: <strong>{fields_str}</strong></p>
            </div>
            """
        
        # Se tutto ok, procedi normalmente
        data = {
            'num_sensors': int(request.form['num_sensors']),
            'device_id': request.form['device_id'],
            'vendor': request.form['vendor'],
            'start_date': request.form['start_date'],
            'start_time': request.form['start_time']
        }
        
        # Formatta la data
        date_obj = datetime.strptime(data['start_date'], "%Y-%m-%d")
        data['start_date_formatted'] = date_obj.strftime("%d-%m-%Y")
        
        # NOVITÀ: Ottiene il vero Device ID dalle API per preview
        print("🔍 Preview: ottenimento Device ID dalle API...")
        transformed_id = transform_device_id_new(data['device_id'])
        metrics = get_sensor_metrics(data['num_sensors'])
        
        start_formatted = format_date_for_sheet(data['start_date'], data['start_time'])
        
        preview_data = {
            'transformed_device_id': transformed_id,
            'metrics': metrics,
            'data': data,
            'start_formatted': start_formatted
        }
        
        return render_template('preview.html', **preview_data)
    
    except Exception as e:
        return f"""
        <div class="alert alert-danger">
            <h5>❌ Errore</h5>
            <p>Si è verificato un errore: <strong>{str(e)}</strong></p>
            <p>Assicurati che tutti i campi siano compilati correttamente e che la connessione di rete sia attiva.</p>
        </div>
        """
    
@app.route('/reset')
def reset_page():
    """Pagina per il reset dei dispositivi"""
    return render_template('reset.html')

@app.route('/reset_device', methods=['POST'])
def reset_device():
    """Esegue il reset del dispositivo con la nuova API semplificata"""
    try:
        device_id = request.form.get('device_id')
        
        if not device_id:
            return render_template('reset_result.html', 
                                 success=False, 
                                 error="Device ID non fornito")
        
        # Parametri di connessione
        ese_test = "device-fabbrica-testing-backend-ese.servizi.prv"
        dev_auth_token = "HDqJXsH41Y84dBE3yyLdfBcpMNI0RGlp"
        
        steps_log = []
        
        # NUOVA API: Una singola chiamata PUT per il reset
        reset_url = f"http://{ese_test}/reset-running-test"
        reset_params = {
            "token": dev_auth_token,
            "clientId": device_id  # ClientID completo es. 1:1:2:15:21:DIGIL_IND_0899
        }
        
        print(f"🔍 Reset dispositivo: {device_id}")
        print(f"📡 Chiamata PUT {reset_url}")
        
        # Esegui la chiamata di reset
        reset_response = requests.put(reset_url, params=reset_params)
        
        steps_log.append(f"PUT /reset-running-test response status: {reset_response.status_code}")
        print(f"PUT /reset-running-test response status: {reset_response.status_code}")
        
        # Verifica il risultato
        if reset_response.status_code == 200:
            print(f"  ✅  RESET SUCCESS per device: {device_id}")
            return render_template('reset_result.html', 
                                 success=True, 
                                 device_id=device_id,
                                 steps_log=steps_log)
        else:
            error_msg = f"Errore reset: HTTP {reset_response.status_code}"
            if reset_response.text:
                error_msg += f" - {reset_response.text[:200]}"
            raise Exception(error_msg)
        
    except requests.exceptions.RequestException as req_err:
        print(f"Errore di rete: {req_err}")
        print(f"  ❌  RESET FAILED")
        return render_template('reset_result.html', 
                             success=False, 
                             device_id=device_id if 'device_id' in locals() else None,
                             error=f"Errore di connessione: {str(req_err)}",
                             steps_log=steps_log if 'steps_log' in locals() else [])
    
    except Exception as e:
        print(f"Exception raised: {e}")
        print(f"  ❌  RESET FAILED")
        return render_template('reset_result.html', 
                             success=False, 
                             device_id=device_id if 'device_id' in locals() else None,
                             error=str(e),
                             steps_log=steps_log if 'steps_log' in locals() else [])
    

@app.route('/test_downlink', methods=['POST'])
def test_downlink():
    """Esegue test Downlink sul dispositivo"""
    try:
        device_id = request.form.get('device_id')
        
        if not device_id:
            return json.dumps({
                'success': False,
                'error': 'Device ID non fornito'
            })
        
        # Trasforma Device ID per le API
        transformed_id = transform_device_id_new(device_id)
        print(f"🔧 Test Downlink: {device_id} -> {transformed_id}")
        
        # Esegui test con ID trasformato
        results = digil_test_service.run_downlink_test(transformed_id)
        
        return json.dumps(results, default=str)
        
    except Exception as e:
        return json.dumps({
            'success': False,
            'error': str(e)
        })
    
@app.route('/test_metrics', methods=['POST'])
def test_metrics():
    """Esegue test Metriche in Range"""
    try:
        device_id = request.form.get('device_id')
        num_sensors = int(request.form.get('num_sensors', 6))
        time_range = int(request.form.get('time_range', 5))
        ui_location = request.form.get('ui_location', 'Lazio')  # Prende UI dal form
        
        if not device_id:
            return json.dumps({
                'success': False,
                'error': 'Device ID non fornito'
            }), 200, {'Content-Type': 'application/json'}
        
        print(f"🔍 Test metriche per device {device_id} con UI {ui_location}")
        
        # Esegui test con UI specificata
        results = digil_test_service.run_metrics_test(
            device_id, 
            num_sensors, 
            ui_location,  # Usa UI dal form
            time_range
        )
        
        return json.dumps(results, default=str), 200, {'Content-Type': 'application/json'}
        
    except Exception as e:
        return json.dumps({
            'success': False,
            'error': str(e)
        }), 200, {'Content-Type': 'application/json'}
    
@app.route('/test_alarm', methods=['POST'])
def test_alarm():
    """Esegue test Allarme Metriche"""
    try:
        device_id = request.form.get('device_id')
        num_sensors = int(request.form.get('num_sensors', 6))
        ui_location = request.form.get('ui_location', 'Lazio')
        time_range = int(request.form.get('time_range', 60))  # Default 60 minuti per allarmi
        
        if not device_id:
            return json.dumps({
                'success': False,
                'error': 'Device ID non fornito'
            }), 200, {'Content-Type': 'application/json'}
        
        print(f"🚨 Test allarmi per device {device_id} con UI {ui_location}")
        print(f"   Range temporale: ultimi {time_range} minuti")
        
        # Esegui test con range temporale
        results = digil_test_service.run_alarm_test(
            device_id, 
            num_sensors, 
            ui_location,
            time_range  # Passa il range temporale
        )
        
        return json.dumps(results, default=str), 200, {'Content-Type': 'application/json'}
        
    except Exception as e:
        return json.dumps({
            'success': False,
            'error': str(e)
        }), 200, {'Content-Type': 'application/json'}

def get_local_ip():
    """Ottiene l'IP locale della macchina"""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"
    
#################### socketio per monitoraggio allarmi ####################

# Aggiungi le route WebSocket
@socketio.on('connect')
def handle_connect():
    print(f'Client connesso: {request.sid}')
    emit('connected', {'data': 'Connesso al server'})

@socketio.on('disconnect')
def handle_disconnect():
    print(f'Client disconnesso: {request.sid}')
    alarm_monitor.stop_monitoring(request.sid)

@socketio.on('start_unified_monitoring')
def handle_start_unified_monitoring(data):
    device_id = data.get('device_id')
    num_sensors = data.get('num_sensors', 6)
    ui = data.get('ui', 'Lazio')
    timeout_minutes = data.get('timeout_minutes', 10)
    
    print(f'Avvio monitoraggio unificato per {device_id}')
    alarm_monitor.start_unified_monitoring(
        request.sid, device_id, num_sensors, ui, timeout_minutes
    )
    emit('monitoring_started', {'message': 'Monitoraggio unificato avviato'})

@socketio.on('stop_monitoring')
def handle_stop_monitoring():
    print(f'Stop monitoraggio per {request.sid}')
    alarm_monitor.stop_monitoring(request.sid)
    emit('monitoring_stopped', {'message': 'Monitoraggio interrotto'})

# Aggiungi route per la pagina di monitoraggio
@app.route('/monitoring')
def monitoring_page():
    return render_template('monitoring.html')

def open_browser():
    """Apre il browser dopo 1.5 secondi"""
    time.sleep(1.5)
    webbrowser.open('http://localhost:5000')

if __name__ == '__main__':
    local_ip = get_local_ip() if 'get_local_ip' in locals() else '127.0.0.1'
    
    if getattr(sys, 'frozen', False):
        threading.Thread(target=open_browser, daemon=True).start()
        print("🚀 DIGIL Report Generator avviato!")
        print("📱 Il browser si aprirà automaticamente...")
        print("🌐 URL locale: http://localhost:5000")
        print(f"🌐 URL rete: http://{local_ip}:5000")
        print("❌ Per chiudere: premi Ctrl+C")
    
    # Usa socketio.run invece di app.run
    socketio.run(app, debug=False, host='0.0.0.0', port=5000)